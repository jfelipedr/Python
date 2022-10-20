from cgi import print_directory
from json import load
from openpyxl import Workbook, load_workbook

wb = load_workbook("C:/Users/INTEL/Desktop/Programacion/Python/Prueba.xlsx")

ws = wb.active
print(ws["B2"].value)

"""wb = Workbook()
ws = wb.active
ws.title = "test1"

ws["A1"] = 123

wb.save("prueba1.xlsx")"""